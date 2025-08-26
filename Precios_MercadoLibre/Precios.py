import requests
import re
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===============================
# FUNCIONES DE SCRAPING
# ===============================

def buscar_precios_ml(producto, limite=10):
    """Busca precios en MercadoLibre Colombia (solo precio al contado)."""
    query = producto.replace(" ", "-")
    url = f"https://listado.mercadolibre.com.co/{query}"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        return []

    precios_raw = re.findall(r'"price"\s*:\s*([\d\.]+)', response.text)
    precios = []
    for p in precios_raw:
        try:
            precios.append(int(p.replace(".", "").strip()))
        except:
            continue

    return precios[:limite]


def calcular_promedios(producto):
    """Calcula promedio de precios en MercadoLibre."""
    precios = buscar_precios_ml(producto)

    if not precios:
        return producto, "No encontrado", "No encontrado", "No encontrado"

    promedio = statistics.mean(precios)
    menos_20 = promedio * 0.8
    menos_30 = promedio * 0.7

    return (
        producto,
        round(promedio),
        round(menos_20),
        round(menos_30)
    )

# ===============================
# FUNCIONES DE EXCEL
# ===============================

def aplicar_estilos(ws):
    """Aplica estilos a la hoja de Excel."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def formato_miles(valor):
    """Formatea número con separador de miles y COP."""
    if isinstance(valor, int) or isinstance(valor, float):
        return f"${valor:,.0f} COP".replace(",", ".")
    return valor

# ===============================
# LISTA DE PRODUCTOS
# ===============================
productos = [
    "Televisor LED 32 pulgadas",
    "Televisor LED 55 pulgadas",
    "Juego de comedor madera 4 puestos",
    "Nevera No Frost",
    "Sala modular",
    "Horno microondas Samsung",
    "Silla ergonómica oficina",
    "Lavadora eléctrica 18 kg",
    "Parrilla eléctrica portátil",
    "Impresora Epson multifuncional",
    "Árbol de navidad con luces",
    "Cafetera Oster",
    "Tostadora eléctrica",
    "Olla a presión eléctrica"
]

# ===============================
# CREAR EXCEL
# ===============================
wb = Workbook()

# Hoja 1: MercadoLibre
ws1 = wb.active
ws1.title = "MercadoLibre"
ws1.append(["Producto", "Promedio", "Promedio -20%", "Promedio -30%"])
for p in productos:
    fila = calcular_promedios(p)
    ws1.append([fila[0], formato_miles(fila[1]), formato_miles(fila[2]), formato_miles(fila[3])])
aplicar_estilos(ws1)

# Guardar archivo
wb.save("precios_productos.xlsx")
print("✅ Archivo 'precios_productos.xlsx' creado con éxito")

